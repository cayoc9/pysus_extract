import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import warnings
import time

# Ignorar avisos do pandas
warnings.filterwarnings('ignore')

def processar_arquivo(arquivo_path, resultados_por_ano):
    """
    Processa um arquivo XLSX e adiciona os resultados ao dicionário por ano.
    
    Args:
        arquivo_path: Caminho do arquivo XLSX
        resultados_por_ano: Dicionário para armazenar resultados por ano
    """
    nome_arquivo = os.path.basename(arquivo_path)
    nome_hospital = os.path.splitext(nome_arquivo)[0]
    
    print(f"\n{'=' * 80}")
    print(f"Processando: {nome_hospital}")
    print(f"{'=' * 80}")
    
    try:
        # Ler o arquivo Excel - cabeçalho na linha 6 (índice 5)
        print(f"Lendo arquivo {arquivo_path}...")
        inicio = time.time()
        df = pd.read_excel(arquivo_path, header=5)
        fim = time.time()
        print(f"Tempo de leitura: {fim - inicio:.2f} segundos")
        print(f"Total de registros: {len(df)}")
        
        # Mostrar informações das colunas encontradas
        print(f"Colunas encontradas: {list(df.columns)}")
        
        # Mapear nomes de colunas (com espaço para com underscore)
        mapeamento_colunas = {
            'Sp Procrea': 'Sp_Procrea',
            'Sp Mm': 'Sp_Mm',
            'Sp Aa': 'Sp_Aa',
            'Sp Cnes': 'Sp_Cnes'
        }
        
        # Renomear colunas se necessário
        colunas_para_renomear = {}
        for coluna_original, coluna_nova in mapeamento_colunas.items():
            if coluna_original in df.columns:
                colunas_para_renomear[coluna_original] = coluna_nova
        
        if colunas_para_renomear:
            print(f"Renomeando colunas: {colunas_para_renomear}")
            df = df.rename(columns=colunas_para_renomear)
            print(f"Colunas após renomeação: {list(df.columns)}")
        
        # Verificar se as colunas necessárias existem
        colunas_necessarias = ['Sp_Procrea', 'Sp_Mm', 'Sp_Aa', 'Sp_Cnes']
        colunas_faltando = [col for col in colunas_necessarias if col not in df.columns]
        
        if colunas_faltando:
            print(f"AVISO: Colunas faltando após renomeação: {colunas_faltando}")
            print(f"Pulando arquivo {nome_hospital}...")
            return
        
        # Verificar dados nulos
        print("Verificando valores nulos por coluna:")
        for col in colunas_necessarias:
            nulos = df[col].isna().sum()
            print(f"  - {col}: {nulos} valores nulos ({nulos/len(df)*100:.2f}%)")
        
        # Para cada ano nos dados
        anos_unicos = df['Sp_Aa'].dropna().unique()
        print(f"Anos encontrados: {anos_unicos}")
        
        for ano in anos_unicos:
            if pd.isna(ano):
                continue
                
            try:
                ano = int(ano)
            except ValueError:
                print(f"AVISO: Valor de ano inválido: {ano}. Pulando...")
                continue
            
            print(f"\nProcessando ano: {ano}")
            
            # Inicializar o dicionário do ano se não existir
            if ano not in resultados_por_ano:
                resultados_por_ano[ano] = {}
            
            # Filtrar dados para o ano atual
            df_ano = df[df['Sp_Aa'] == ano].copy()
            print(f"  - Registros para o ano {ano}: {len(df_ano)}")
            
            # Para cada procedimento único, contar por mês
            procedimentos = df_ano['Sp_Procrea'].dropna().unique()
            print(f"  - Procedimentos únicos: {len(procedimentos)}")
            
            if len(procedimentos) == 0:
                print(f"AVISO: Nenhum procedimento encontrado para o ano {ano}. Pulando...")
                continue
            
            # Amostra dos procedimentos encontrados
            amostra = procedimentos[:5] if len(procedimentos) > 5 else procedimentos
            print(f"  - Amostra de procedimentos: {amostra}")
            
            # Dataframe para armazenar os resultados
            dados = []
            
            # Contador para progresso
            contador = 0
            total_procedimentos = len(procedimentos)
            print(f"Iniciando processamento de {total_procedimentos} procedimentos...")
            
            inicio_proc = time.time()
            for proc in procedimentos:
                # Atualização de progresso a cada 100 procedimentos
                contador += 1
                if contador % 100 == 0 or contador == total_procedimentos:
                    print(f"  - Progresso: {contador}/{total_procedimentos} ({contador/total_procedimentos*100:.1f}%)")
                
                # Linha para o procedimento atual
                linha = {'Procedimento': proc}
                
                # Adicionar grupo (primeiros 2 dígitos)
                try:
                    linha['Grupo'] = str(proc)[:2]
                except:
                    print(f"AVISO: Erro ao processar grupo para procedimento {proc}. Usando '00'")
                    linha['Grupo'] = '00'
                
                # Calcular contagem para cada mês
                total = 0
                for mes in range(1, 13):
                    count = len(df_ano[(df_ano['Sp_Procrea'] == proc) & (df_ano['Sp_Mm'] == mes)])
                    linha[f'Mês {mes}'] = count
                    total += count
                
                linha['Total'] = total
                dados.append(linha)
            
            fim_proc = time.time()
            print(f"Tempo de processamento: {fim_proc - inicio_proc:.2f} segundos")
            
            # Criar DataFrame com os resultados
            if dados:
                resultado_df = pd.DataFrame(dados)
                
                # Ordenar por grupo e procedimento
                resultado_df = resultado_df.sort_values(['Grupo', 'Procedimento'])
                
                # Reordenar as colunas
                colunas = ['Procedimento'] + [f'Mês {mes}' for mes in range(1, 13)] + ['Total', 'Grupo']
                resultado_df = resultado_df[colunas]
                
                # Mostrar estatísticas do resultado
                print(f"Resultado final:")
                print(f"  - Total de procedimentos: {len(resultado_df)}")
                print(f"  - Grupos de procedimentos: {resultado_df['Grupo'].nunique()}")
                print(f"  - Total de procedimentos realizados: {resultado_df['Total'].sum()}")
                
                # Adicionar ao dicionário de resultados
                resultados_por_ano[ano][nome_hospital] = resultado_df
                
                print(f"Processamento do ano {ano} concluído com sucesso!")
            else:
                print(f"AVISO: Nenhum resultado gerado para o ano {ano}.")
            
    except Exception as e:
        print(f"ERRO ao processar arquivo {nome_hospital}: {str(e)}")
        import traceback
        print(traceback.format_exc())

def formatar_planilha(workbook, sheet_name):
    """
    Formata uma planilha Excel com cores alternadas para os grupos.
    
    Args:
        workbook: Workbook de openpyxl
        sheet_name: Nome da planilha
    """
    print(f"Formatando planilha: {sheet_name}")
    worksheet = workbook[sheet_name]
    
    # Formatar cabeçalho
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    
    # Aplicar cores alternadas aos grupos
    grupo_atual = None
    cor_idx = 0
    cores = [
        PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    ]
    
    # Começar da linha 2 (após o cabeçalho)
    for row_idx in range(2, worksheet.max_row + 1):
        # Obter o grupo da última coluna
        grupo = worksheet.cell(row=row_idx, column=worksheet.max_column).value
        
        if grupo != grupo_atual:
            grupo_atual = grupo
            cor_idx = (cor_idx + 1) % 2
        
        # Aplicar cor a todas as células exceto a última (Grupo)
        for col_idx in range(1, worksheet.max_column):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = cores[cor_idx]
    
    # Ajustar largura das colunas
    for col_idx in range(1, worksheet.max_column + 1):
        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
        max_length = 0
        
        for row_idx in range(1, min(100, worksheet.max_row + 1)):  # Limitar a 100 linhas para performance
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        worksheet.column_dimensions[col_letter].width = max_length + 2
    
    # Ocultar a coluna de Grupo (última coluna)
    last_col_letter = worksheet.cell(row=1, column=worksheet.max_column).column_letter
    worksheet.column_dimensions[last_col_letter].hidden = True
    
    print(f"Formatação de {sheet_name} concluída!")

def salvar_resultados(resultados_por_ano):
    """
    Salva os resultados em arquivos Excel por ano.
    
    Args:
        resultados_por_ano: Dicionário com os resultados por ano
    """
    if not resultados_por_ano:
        print("AVISO: Não há resultados para salvar!")
        return
    
    for ano, hospitais in resultados_por_ano.items():
        if not hospitais:
            print(f"AVISO: Não há hospitais com dados para o ano {ano}. Pulando...")
            continue
        
        nome_arquivo = f"Procedimentos Aprovados {ano}.xlsx"
        print(f"\nSalvando {nome_arquivo}...")
        
        # Criar arquivo Excel com pandas
        inicio = time.time()
        with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
            for hospital, df in hospitais.items():
                # Limitar nome da planilha a 31 caracteres (limite do Excel)
                nome_planilha = hospital[:31]
                print(f"  - Salvando planilha '{nome_planilha}' com {len(df)} linhas")
                df.to_excel(writer, sheet_name=nome_planilha, index=False)
        
        # Aplicar formatação adicional
        print("Aplicando formatação...")
        workbook = load_workbook(nome_arquivo)
        
        for hospital in hospitais:
            nome_planilha = hospital[:31]
            formatar_planilha(workbook, nome_planilha)
        
        workbook.save(nome_arquivo)
        fim = time.time()
        print(f"Arquivo {nome_arquivo} salvo com sucesso! Tempo total: {fim - inicio:.2f} segundos")

def main():
    # Caminho da pasta com os arquivos
    pasta_dados = "/home/cayo/repos/ICSF/Pysus/pysus_extract/Analises/relatorios/procedimentos_aprovados_ES"
    
    # Verificar se a pasta existe
    if not os.path.exists(pasta_dados):
        print(f"Pasta {pasta_dados} não encontrada!")
        return
    
    print(f"Pasta encontrada: {pasta_dados}")
    
    # Listar todos os arquivos XLSX na pasta
    arquivos = [os.path.join(pasta_dados, f) for f in os.listdir(pasta_dados) 
                if f.endswith('.xlsx') and os.path.isfile(os.path.join(pasta_dados, f))
                and not f.startswith('.~lock')]  # Ignorar arquivos temporários do LibreOffice
    
    print(f"Encontrados {len(arquivos)} arquivos para processar.")
    
    # Dicionário para armazenar resultados por ano
    resultados_por_ano = {}
    
    # Processar cada arquivo
    tempo_inicio_geral = time.time()
    for i, arquivo in enumerate(arquivos):
        print(f"\nProcessando arquivo {i+1} de {len(arquivos)}: {os.path.basename(arquivo)}")
        processar_arquivo(arquivo, resultados_por_ano)
    
    # Salvar os resultados
    salvar_resultados(resultados_por_ano)
    
    tempo_fim_geral = time.time()
    print(f"\nProcessamento concluído! Tempo total: {(tempo_fim_geral - tempo_inicio_geral) / 60:.2f} minutos")

if __name__ == "__main__":
    main() 