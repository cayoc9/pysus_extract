import os
import pandas as pd
import duckdb
import logging
import time
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Tuple
import warnings
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import gc

# Configurações globais
warnings.filterwarnings('ignore')
logging.basicConfig(level=logging.INFO)

def carregar_dados_parquet(base: str, grupo: str, cnes_list: List[str], 
                         competencia_inicio: str, competencia_fim: str) -> pd.DataFrame:
    """
    Carrega dados dos arquivos parquet usando as funções de main.py
    
    Args:
        base: SIH (Sistema de Informações Hospitalares)
        grupo: Código do grupo (RJ ou RD)
        cnes_list: Lista de CNES ou ["*"] para todos
        competencia_inicio: Mês/Ano inicial (MM/YYYY)
        competencia_fim: Mês/Ano final (MM/YYYY)
    
    Returns:
        DataFrame com os dados carregados
    """
    from main import get_parquet_files, QueryParams, process_data
    
    logging.info(f"Carregando dados para {base}/{grupo} de {competencia_inicio} a {competencia_fim}")
    
    try:
        # Lista de campos genéricos sem prefixos para evitar erros de binding
        # Importante: os nomes devem corresponder às colunas reais nos arquivos
        campos_especificos = [
            # Campos principais para a análise
            "GESTAO", "UF_ZI", "ANO_CMPT", "MES_CMPT", "CNES", 
            "N_AIH", "DT_INTER", "DT_SAIDA", "NUM_PROC",
            "VALOR_TOT", "VALOR_SH", "PROC_REA"
        ]
        
        # Define os parâmetros da consulta
        params = QueryParams(
            base=base,
            grupo=grupo,
            cnes_list=cnes_list,
            campos_agrupamento=campos_especificos,
            competencia_inicio=competencia_inicio,
            competencia_fim=competencia_fim,
            table_name=None,
            consulta_personalizada=None
        )
        
        # Obtém os arquivos parquet
        files = get_parquet_files(base, grupo, competencia_inicio, competencia_fim)
        if not files:
            logging.warning(f"Nenhum arquivo encontrado para {base}/{grupo}")
            return pd.DataFrame()
        
        logging.info(f"Encontrados {len(files)} arquivos para processamento")
        
        # Processa os dados
        temp_table = process_data(files, params)
        logging.info(f"Tabela temporária criada: {temp_table}")
        
        # Converter para DataFrame 
        df = duckdb.query(f"SELECT * FROM {temp_table}").to_df()
        logging.info(f"Colunas disponíveis: {', '.join(df.columns)}")
        
        # Verificação de registros
        if len(df) > 0:
            try:
                # Encontrar a coluna de AIH (pode ser N_AIH ou outra variação)
                aih_col = None
                for col in df.columns:
                    if 'AIH' in col.upper() or 'NAIH' in col.upper():
                        aih_col = col
                        break
                
                if aih_col:
                    contagem_aih = df[aih_col].value_counts().head(10)
                    logging.info(f"Top 10 AIH na tabela: \n{contagem_aih}")
                    
                    # Adicionar prefixo ao grupo para identificar origem
                    df = df.add_prefix(f"{grupo}_")
                    
                else:
                    logging.warning("Coluna de AIH não encontrada nos dados")
            except Exception as e:
                logging.warning(f"Erro ao analisar contagem por AIH: {e}")
        
        logging.info(f"Dados carregados: {len(df)} registros com {len(df.columns)} colunas")
        return df
        
    except Exception as e:
        logging.error(f"Erro ao carregar dados: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return pd.DataFrame()

def padronizar_data(data_str: str) -> datetime:
    """
    Padroniza o formato de data para objeto datetime
    
    Args:
        data_str: String com a data no formato YYYYMMDD
        
    Returns:
        Objeto datetime
    """
    if pd.isna(data_str) or data_str == '':
        return None
    
    try:
        # Garantir formato YYYYMMDD
        data_str = str(data_str).strip()
        
        # Para caso o formato seja YYYYMMDD
        if len(data_str) == 8:
            return datetime.strptime(data_str, '%Y%m%d')
        
        # Caso tenha separadores (YYYY-MM-DD ou YYYY/MM/DD)
        if '-' in data_str or '/' in data_str:
            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%d-%m-%Y']:
                try:
                    return datetime.strptime(data_str, fmt)
                except:
                    continue
                    
        logging.warning(f"Formato de data não reconhecido: {data_str}")
        return None
    except Exception as e:
        logging.warning(f"Erro ao converter data {data_str}: {str(e)}")
        return None

def analisar_perdas_reais(df_rejeitadas: pd.DataFrame, df_aprovadas: pd.DataFrame) -> pd.DataFrame:
    """
    Analisa as perdas reais comparando AIHs rejeitadas vs. aprovadas
    
    Args:
        df_rejeitadas: DataFrame com AIHs rejeitadas (grupo RJ)
        df_aprovadas: DataFrame com AIHs aprovadas (grupo RD)
    
    Returns:
        DataFrame com análise de perdas reais
    """
    logging.info("Iniciando análise de perdas reais")
    
    # Verificar quais colunas estão disponíveis
    logging.info(f"Colunas RJ: {df_rejeitadas.columns.tolist()}")
    logging.info(f"Colunas RD: {df_aprovadas.columns.tolist()}")
    
    # Identificar colunas para AIH e DATA_SAIDA nos dois DataFrames
    # Para RJ (rejeitadas)
    col_naih_rj = None
    col_dtsaida_rj = None
    for col in df_rejeitadas.columns:
        if 'AIH' in col.upper() or 'NAIH' in col.upper():
            col_naih_rj = col
        if 'SAIDA' in col.upper() or 'DTSAIDA' in col.upper():
            col_dtsaida_rj = col
    
    # Para RD (aprovadas)
    col_naih_rd = None
    col_dtsaida_rd = None
    for col in df_aprovadas.columns:
        if 'AIH' in col.upper() or 'NAIH' in col.upper():
            col_naih_rd = col
        if 'SAIDA' in col.upper() or 'DTSAIDA' in col.upper():
            col_dtsaida_rd = col
    
    # Verificar se encontramos as colunas necessárias
    if not all([col_naih_rj, col_dtsaida_rj, col_naih_rd, col_dtsaida_rd]):
        logging.error(f"Colunas necessárias não encontradas: NAIH_RJ={col_naih_rj}, DTSAIDA_RJ={col_dtsaida_rj}, NAIH_RD={col_naih_rd}, DTSAIDA_RD={col_dtsaida_rd}")
        return pd.DataFrame()
    
    logging.info(f"Usando colunas: NAIH_RJ={col_naih_rj}, DTSAIDA_RJ={col_dtsaida_rj}, NAIH_RD={col_naih_rd}, DTSAIDA_RD={col_dtsaida_rd}")
    
    # Padronizando os códigos de AIH
    df_rejeitadas['AIH_NORMALIZADA'] = df_rejeitadas[col_naih_rj].astype(str).str.strip()
    df_aprovadas['AIH_NORMALIZADA'] = df_aprovadas[col_naih_rd].astype(str).str.strip()
    
    # Convertendo datas de saída para datetime
    df_rejeitadas['DATA_SAIDA'] = df_rejeitadas[col_dtsaida_rj].apply(padronizar_data)
    df_aprovadas['DATA_SAIDA'] = df_aprovadas[col_dtsaida_rd].apply(padronizar_data)
    
    # Remover registros com data de saída inválida
    df_rejeitadas = df_rejeitadas.dropna(subset=['DATA_SAIDA'])
    
    # Criar um dicionário de AIHs aprovadas para busca rápida
    aihs_aprovadas = set(df_aprovadas['AIH_NORMALIZADA'].tolist())
    
    # Preparar DataFrame para resultado
    resultado = []
    
    # Processar cada AIH rejeitada
    for _, row in df_rejeitadas.iterrows():
        aih = row['AIH_NORMALIZADA']
        data_saida = row['DATA_SAIDA']
        
        # Calcular prazo limite (6 meses após data de saída)
        prazo_limite = data_saida + timedelta(days=180)
        
        # Verificar se esta AIH aparece no conjunto de aprovadas
        encontrada = aih in aihs_aprovadas
        
        # Se encontrada, verificar se está dentro do prazo
        dentro_prazo = None
        data_aprovacao = None
        
        if encontrada:
            # Encontrar todas as ocorrências desta AIH nas aprovadas
            aprovacoes = df_aprovadas[df_aprovadas['AIH_NORMALIZADA'] == aih]
            
            for _, apr_row in aprovacoes.iterrows():
                data_apr = apr_row['DATA_SAIDA']
                if data_apr and data_apr <= prazo_limite:
                    dentro_prazo = True
                    data_aprovacao = data_apr
                    break
            
            if dentro_prazo is None:
                dentro_prazo = False
        else:
            dentro_prazo = False
        
        # É perda real se não foi encontrada ou se foi encontrada fora do prazo
        perda_real = not (encontrada and dentro_prazo)
        
        # Adicionar ao resultado
        resultado.append({
            'AIH': aih,
            'Data_Saida': data_saida,
            'Prazo_Limite': prazo_limite,
            'Encontrada': encontrada,
            'Dentro_Prazo': dentro_prazo,
            'Data_Aprovacao': data_aprovacao,
            'Perda_Real': perda_real
        })
    
    # Criar DataFrame de resultado
    resultado_df = pd.DataFrame(resultado)
    
    # Adicionar estatísticas
    total_rejeitadas = len(resultado_df)
    total_perdas = resultado_df['Perda_Real'].sum()
    percentual_perda = (total_perdas / total_rejeitadas * 100) if total_rejeitadas > 0 else 0
    
    logging.info(f"Análise concluída: {total_rejeitadas} AIHs rejeitadas, {total_perdas} perdas reais ({percentual_perda:.2f}%)")
    
    return resultado_df

def formatar_planilha(workbook, sheet_name, cnes, periodo_inicio, periodo_fim, uf="SP"):
    """
    Formata a planilha Excel com o layout desejado para o relatório de perdas reais
    """
    worksheet = workbook[sheet_name]
    
    # Adicionar cabeçalho
    worksheet.insert_rows(1, 4)  # Insere 4 linhas no topo
    
    # Linha 1: Título principal
    worksheet.cell(row=1, column=1, value=f"Análise de Perdas Reais de AIH - {uf}")
    # Linha 2: Subtítulo
    worksheet.cell(row=2, column=1, value="AIHs Rejeitadas vs. Aprovadas")
    # Linha 3: Estabelecimento
    if isinstance(cnes, list) and len(cnes) == 1:
        worksheet.cell(row=3, column=1, value=f"Estabelecimento: '{cnes[0]}'")
    else:
        worksheet.cell(row=3, column=1, value=f"Estabelecimentos: Múltiplos")
    # Linha 4: Período
    worksheet.cell(row=4, column=1, value=f"Periodo: {periodo_inicio}-{periodo_fim}")
    
    # Aplicar formatação ao cabeçalho
    header_font = Font(bold=True, size=12)
    for row in range(1, 5):
        cell = worksheet.cell(row=row, column=1)
        cell.font = header_font
        # Mesclar células nas linhas de cabeçalho
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    
    # Formatar cabeçalho de colunas
    header_cells = [
        'AIH', 'Data de Saída', 'Prazo Limite', 'Encontrada', 
        'Dentro do Prazo', 'Data de Aprovação', 'Perda Real'
    ]
    
    for col_idx, header in enumerate(header_cells, 1):
        cell = worksheet.cell(row=5, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Formatar todas as linhas de dados
    for row_idx in range(6, worksheet.max_row + 1):
        # Alternar cores de fundo
        fill = PatternFill(start_color="F2F2F2" if row_idx % 2 == 0 else "E6E6E6", 
                           end_color="F2F2F2" if row_idx % 2 == 0 else "E6E6E6", 
                           fill_type="solid")
        
        for col_idx in range(1, 8):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            
            # Formatar colunas específicas
            if col_idx in [4, 5, 7]:  # Colunas booleanas
                valor = cell.value
                if valor is not None:
                    # Converter para "Sim" ou "Não"
                    cell.value = "Sim" if valor else "Não"
                    # Destacar perdas reais em vermelho
                    if col_idx == 7 and valor:
                        cell.font = Font(color="FF0000", bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Centralizar colunas de data
            elif col_idx in [2, 3, 6]:
                cell.alignment = Alignment(horizontal='center')
                # Formatar datas se necessário
                if isinstance(cell.value, datetime):
                    cell.number_format = 'dd/mm/yyyy'
    
    # Ajustar largura das colunas
    worksheet.column_dimensions['A'].width = 20  # AIH
    worksheet.column_dimensions['B'].width = 15  # Data Saída
    worksheet.column_dimensions['C'].width = 15  # Prazo Limite
    worksheet.column_dimensions['D'].width = 15  # Encontrada
    worksheet.column_dimensions['E'].width = 15  # Dentro Prazo
    worksheet.column_dimensions['F'].width = 15  # Data Aprovação
    worksheet.column_dimensions['G'].width = 15  # Perda Real
    
    logging.info(f"Planilha {sheet_name} formatada com sucesso")

def gerar_relatorio_perdas_reais(base: str, cnes_list: List[str],
                                competencia_inicio: str, competencia_fim: str,
                                output_dir: str = "relatorios"):
    """
    Gera relatório de perdas reais de AIHs
    
    Args:
        base: SIH 
        cnes_list: Lista de CNES ou ["*"] para todos
        competencia_inicio: Mês/Ano inicial (MM/YYYY)
        competencia_fim: Mês/Ano final (MM/YYYY)
        output_dir: Diretório para salvar os arquivos de saída
    """
    inicio_total = time.time()
    os.makedirs(output_dir, exist_ok=True)
    
    try:
        # Carregar dados rejeitados (RJ)
        df_rejeitadas = carregar_dados_parquet(
            base=base,
            grupo="RJ",
            cnes_list=cnes_list,
            competencia_inicio=competencia_inicio,
            competencia_fim=competencia_fim
        )
        
        if df_rejeitadas.empty:
            logging.error("Não foi possível carregar dados de AIHs rejeitadas (RJ)")
            return
        
        # Carregar dados aprovados (RD)
        df_aprovadas = carregar_dados_parquet(
            base=base,
            grupo="RD",
            cnes_list=cnes_list,
            competencia_inicio=competencia_inicio,
            competencia_fim=competencia_fim
        )
        
        if df_aprovadas.empty:
            logging.error("Não foi possível carregar dados de AIHs aprovadas (RD)")
            return
        
        # Analisar perdas reais
        resultado_df = analisar_perdas_reais(df_rejeitadas, df_aprovadas)
        
        if resultado_df.empty:
            logging.error("Não foi possível analisar perdas reais - resultado vazio")
            return
        
        # Nome do arquivo de saída
        nome_arquivo = os.path.join(output_dir, f"Perdas_Reais_AIH_{competencia_inicio.replace('/', '')}_a_{competencia_fim.replace('/', '')}.xlsx")
        
        # Criar relatório Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Perdas Reais"
        
        # Adicionar dados ao Excel
        headers = ['AIH', 'Data_Saida', 'Prazo_Limite', 'Encontrada', 'Dentro_Prazo', 'Data_Aprovacao', 'Perda_Real']
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        for row_idx, row in enumerate(resultado_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Formatar planilha
        formatar_planilha(wb, "Perdas Reais", cnes_list, competencia_inicio, competencia_fim)
        
        # Adicionar resumo estatístico
        ws_resumo = wb.create_sheet("Resumo")
        
        # Estatísticas básicas
        total_rejeitadas = len(resultado_df)
        total_perdas = resultado_df['Perda_Real'].sum()
        percentual_perda = (total_perdas / total_rejeitadas * 100) if total_rejeitadas > 0 else 0
        
        # Adicionar estatísticas ao resumo
        ws_resumo.cell(row=1, column=1, value="Estatísticas de Perdas Reais")
        ws_resumo.cell(row=2, column=1, value="Total de AIHs Rejeitadas:")
        ws_resumo.cell(row=2, column=2, value=total_rejeitadas)
        ws_resumo.cell(row=3, column=1, value="Total de Perdas Reais:")
        ws_resumo.cell(row=3, column=2, value=total_perdas)
        ws_resumo.cell(row=4, column=1, value="Percentual de Perda:")
        ws_resumo.cell(row=4, column=2, value=f"{percentual_perda:.2f}%")
        
        # Formatar aba de resumo
        for row in range(1, 5):
            ws_resumo.cell(row=row, column=1).font = Font(bold=True)
            
        ws_resumo.column_dimensions['A'].width = 25
        ws_resumo.column_dimensions['B'].width = 15
        
        # Salvar arquivo
        wb.save(nome_arquivo)
        logging.info(f"Relatório salvo com sucesso em {nome_arquivo}")
        
        # Liberar memória
        del df_rejeitadas
        del df_aprovadas
        del resultado_df
        gc.collect()
        
        logging.info(f"Processamento concluído em {time.time() - inicio_total:.2f} segundos")
        
    except Exception as e:
        logging.error(f"Erro durante a geração do relatório: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())

def main():
    """
    Função principal - exemplo de uso
    """
    # Configurações
    base = "SIH"
    cnes_list = [
        "0011800",
        "0012017",
        "2445956",
        "2446030",
        "2447894",
        "2465752",
        "2486199"
    ]  # Lista de CNES ou ["*"] para todos
    competencia_inicio = "01/2024"
    competencia_fim = "05/2024"
    output_dir = "relatorios/perdas_reais"
    
    # Gerar relatório
    gerar_relatorio_perdas_reais(
        base=base,
        cnes_list=cnes_list,
        competencia_inicio=competencia_inicio,
        competencia_fim=competencia_fim,
        output_dir=output_dir
    )

if __name__ == "__main__":
    main()