import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
import time
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
import duckdb
import numpy as np
from collections import defaultdict

# Importações do projeto - remover importações diretas para evitar circular import
from api.definitions import procedimentos_dict, CAMPOS_CNES, GRUPOS_INFO, MAPEAMENTO_CNES
# Importando funções de utilidade do módulo utils
from api.utils import get_parquet_files, process_data
from api.definitions import QueryParams

# Configurações globais
warnings.filterwarnings('ignore')
logging.basicConfig(level=logging.INFO)

def carregar_dados_parquet(base: str, grupo: str, cnes_list: List[str], 
                          competencia_inicio: str, competencia_fim: str) -> pd.DataFrame:
    """
    Carrega dados dos arquivos parquet usando as funções de api.utils
    
    Args:
        base: SIH ou SIA
        grupo: Código do grupo (SP, RD, etc)
        cnes_list: Lista de CNES ou ["*"] para todos
        competencia_inicio: Mês/Ano inicial (MM/YYYY)
        competencia_fim: Mês/Ano final (MM/YYYY)
    
    Returns:
        DataFrame com os dados carregados
    """
    # Usando as funções importadas de api.utils
    logging.info(f"Carregando dados para {base}/{grupo} de {competencia_inicio} a {competencia_fim}")
    
    try:
        # Lista de campos específicos para SP - campos importantes para análise de procedimentos
        campos_especificos = [
            "SP_GESTOR", "SP_UF", "SP_AA", "SP_MM", "SP_CNES", "SP_NAIH", 
            "SP_PROCREA", "SP_DTINTER", "SP_DTSAIDA", "SP_NUM_PR", 
            "SP_TIPO", "SP_CPFCGC", "SP_ATOPROF", "SP_TP_ATO", "SP_QTD_ATO"
        ]
        
        # Define os parâmetros da consulta
        params = QueryParams(
            base=base,
            grupo=grupo,
            cnes_list=cnes_list,
            campos_agrupamento=campos_especificos,
            competencia_inicio=competencia_inicio,
            competencia_fim=competencia_fim,
            table_name=None,
            consulta_personalizada=None
        )
        
        # Obtém os arquivos parquet
        files = get_parquet_files(base, grupo, competencia_inicio, competencia_fim)
        if not files:
            logging.warning(f"Nenhum arquivo encontrado para {base}/{grupo}")
            return pd.DataFrame()
        
        logging.info(f"Encontrados {len(files)} arquivos para processamento")
        
        # Usar o método process_data que funcionou anteriormente
        temp_table = process_data(files, params)
        logging.info(f"Tabela temporária criada: {temp_table}")
        
        # Converter para DataFrame e verificar os resultados
        df = duckdb.query(f"SELECT * FROM {temp_table}").to_df()
        logging.info(f"Colunas disponíveis: {', '.join(df.columns)}")
        
        # Contar registros por CNES para verificar se os dados estão corretos
        if len(df) > 0:
            try:
                # Encontrar a coluna CNES - pode ter várias grafias
                cnes_col = None
                for col in df.columns:
                    if 'CNES' in col.upper():
                        cnes_col = col
                        break
                
                if cnes_col:
                    contagem_cnes = df[cnes_col].value_counts().head(10)
                    logging.info(f"Top 10 CNES na tabela: \n{contagem_cnes}")
            except Exception as e:
                logging.warning(f"Erro ao analisar contagem por CNES: {e}")
        
        logging.info(f"Dados carregados: {len(df)} registros com {len(df.columns)} colunas")
        return df
        
    except Exception as e:
        logging.error(f"Erro ao carregar dados: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return pd.DataFrame()

def normalizar_procedimento(codigo: str) -> str:
    """
    Garante que o código do procedimento tenha 10 dígitos, preservando zeros à esquerda
    """
    if not codigo or pd.isna(codigo):
        return ""
    
    # Converte para string se não for
    codigo_str = str(codigo)
    
    # Remove espaços, pontos e outros caracteres
    codigo_limpo = ''.join(c for c in codigo_str if c.isdigit())
    
    # Garante que tenha 10 dígitos, preenchendo com zeros à esquerda se necessário
    return codigo_limpo.zfill(10)

def obter_nome_procedimento(codigo: str) -> str:
    """
    Obtém o nome do procedimento a partir do código usando o dicionário procedimentos_dict
    """
    codigo_normalizado = normalizar_procedimento(codigo)
    return procedimentos_dict.get(codigo_normalizado, "Procedimento não identificado")

def obter_nome_hospital(cnes: str) -> str:
    """
    Obtém o nome do hospital a partir do CNES usando o mapeamento de definitions.py
    
    Args:
        cnes: Código CNES do estabelecimento de saúde (pode ser string ou número)
        
    Returns:
        str: Nome do hospital ou código CNES se não encontrado
    """
    try:
        # Garantir que o CNES seja string
        if not isinstance(cnes, str):
            cnes = str(cnes)
        
        # Remover espaços e caracteres não numéricos
        cnes_limpo = ''.join(c for c in cnes if c.isdigit())
        
        # Garantir que o CNES tenha 7 dígitos, preenchendo com zeros à esquerda
        if len(cnes_limpo) <= 7:
            cnes_limpo = cnes_limpo.zfill(7)
        else:
            # Se tiver mais de 7 dígitos, usar apenas os últimos 7
            cnes_limpo = cnes_limpo[-7:]
            logging.warning(f"CNES possui mais de 7 dígitos, usando apenas os últimos 7: {cnes_limpo}")
        
        # Busca no mapeamento importado de definitions.py
        if hasattr(MAPEAMENTO_CNES, 'get'):
            return MAPEAMENTO_CNES.get(cnes_limpo, f"Hospital {cnes_limpo}")
        else:
            logging.warning("MAPEAMENTO_CNES não é um dicionário. Usando código CNES como nome.")
            return f"Hospital {cnes_limpo}"
            
    except Exception as e:
        logging.warning(f"Erro ao processar CNES {cnes}: {str(e)}")
        return f"Hospital {cnes}"

def analisar_procedimentos_hospital(df: pd.DataFrame, ano: int) -> pd.DataFrame:
    """
    Analisa os procedimentos de um hospital em um ano específico, agrupando tanto procedimentos
    principais (SP_PROCREA) quanto secundários (SP_ATOPROF)
    
    Args:
        df: DataFrame com os dados do hospital
        ano: Ano de referência
    
    Returns:
        DataFrame com a análise de procedimentos por mês
    """
    logging.info(f"Analisando procedimentos para o ano {ano}")
    
    # Verificar quais colunas estão disponíveis
    logging.info(f"Colunas disponíveis: {df.columns.tolist()}")
    
    # Garantir que existam colunas de ano e mês
    col_ano = None
    if 'SP_AA' in df.columns:
        col_ano = 'SP_AA'
    elif 'ANO_CMPT' in df.columns:
        col_ano = 'ANO_CMPT'
    else:
        # Procurar qualquer coluna que tenha 'ANO' ou 'AA' no nome
        for col in df.columns:
            if 'ANO' in col.upper() or 'AA' in col.upper():
                col_ano = col
                break
    
    # Se não encontrou coluna de ano, não podemos prosseguir
    if not col_ano:
        logging.error("Não foi possível encontrar coluna de ano nos dados")
        return pd.DataFrame()
    
    logging.info(f"Usando coluna {col_ano} para o ano")
    
    # Converter coluna de ano para inteiro se necessário
    try:
        if df[col_ano].dtype != 'int64':
            df[col_ano] = df[col_ano].astype(int)
    except Exception as e:
        logging.warning(f"Erro ao converter coluna de ano para inteiro: {e}")
    
    # Filtrar dados por ano
    df_ano = df[df[col_ano] == ano].copy()
    
    # Identificar coluna de mês
    col_mes = None
    if 'SP_MM' in df.columns:
        col_mes = 'SP_MM'
    elif 'MES_CMPT' in df.columns:
        col_mes = 'MES_CMPT'
    else:
        # Procurar qualquer coluna que tenha 'MES' ou 'MM' no nome
        for col in df.columns:
            if 'MES' in col.upper() or 'MM' in col.upper():
                col_mes = col
                break
    
    # Se não encontrou coluna de mês, não podemos prosseguir
    if not col_mes:
        logging.error("Não foi possível encontrar coluna de mês nos dados")
        return pd.DataFrame()
    
    logging.info(f"Usando coluna {col_mes} para o mês")
    
    # Converter coluna de mês para inteiro se necessário
    try:
        if df_ano[col_mes].dtype != 'int64':
            df_ano[col_mes] = df_ano[col_mes].astype(int)
    except Exception as e:
        logging.warning(f"Erro ao converter coluna de mês para inteiro: {e}")
    
    # Se não houver dados para o ano, retornar DataFrame vazio
    if len(df_ano) == 0:
        logging.warning(f"Nenhum dado encontrado para o ano {ano}")
        return pd.DataFrame()
    else:
        logging.info(f"Encontrados {len(df_ano)} registros para o ano {ano}")
        
        # Mostrar distribuição por mês para debug
        mes_counts = df_ano[col_mes].value_counts().sort_index()
        logging.info(f"Distribuição por mês:\n{mes_counts}")
    
    # Identificar colunas de procedimentos
    colunas_proc = []
    if 'SP_PROCREA' in df_ano.columns:
        colunas_proc.append('SP_PROCREA')
    if 'SP_ATOPROF' in df_ano.columns:
        colunas_proc.append('SP_ATOPROF')
    
    # Se não encontrar as colunas específicas, procurar alternativas
    if not colunas_proc:
        possiveis_colunas = [col for col in df_ano.columns if 'PROC' in col.upper()]
        colunas_proc.extend(possiveis_colunas)
    
    if not colunas_proc:
        logging.error("Nenhuma coluna de procedimento encontrada")
        return pd.DataFrame()
    
    logging.info(f"Usando as colunas {colunas_proc} para códigos de procedimento")
    
    # Normalizar códigos de procedimentos (preservando zeros à esquerda)
    for col in colunas_proc:
        df_ano[col] = df_ano[col].apply(normalizar_procedimento)
    
    # Coletar todos os procedimentos únicos de todas as colunas
    todos_procedimentos = set()
    for col in colunas_proc:
        todos_procedimentos.update(df_ano[col].dropna().unique())
    
    # Remover strings vazias se existirem
    if '' in todos_procedimentos:
        todos_procedimentos.remove('')
    
    logging.info(f"Encontrados {len(todos_procedimentos)} procedimentos únicos em todas as colunas")
    
    # Criar DataFrame de resultado
    dados = []
    
    for i, proc in enumerate(sorted(todos_procedimentos)):
        if i % 100 == 0:
            logging.info(f"Processando procedimento {i+1}/{len(todos_procedimentos)}")
        
        # Obter grupo do procedimento (2 primeiros dígitos)
        grupo = proc[:2] if len(proc) >= 2 else "00"
        
        # Obter nome do procedimento
        nome_proc = obter_nome_procedimento(proc)
        
        # Nome completo para exibição
        nome_completo = f"{proc} - {nome_proc}"
        
        # Inicializar linha com o procedimento
        linha = {
            'Procedimento': nome_completo,
            'Grupo': grupo
        }
        
        # Calcular contagem por mês para todas as colunas de procedimento combinadas
        total = 0
        for mes in range(1, 13):
            # Condição de filtro para o mês
            condicao_mes = df_ano[col_mes] == mes
            
            # Inicializa contagem para o mês
            count = 0
            
            # Soma ocorrências em todas as colunas de procedimento
            for col in colunas_proc:
                count_col = len(df_ano[condicao_mes & (df_ano[col] == proc)])
                count += count_col
            
            linha[f'Mês {mes}'] = count
            total += count
        
        linha['Total'] = total
        
        # Somente adicionar procedimentos com pelo menos uma ocorrência
        if total > 0:
            dados.append(linha)
    
    # Criar DataFrame e ordenar por grupo e procedimento
    if dados:
        resultado_df = pd.DataFrame(dados)
        resultado_df = resultado_df.sort_values(['Grupo', 'Procedimento'])
        
        # Reorganizar colunas
        colunas = ['Procedimento'] + [f'Mês {mes}' for mes in range(1, 13)] + ['Total', 'Grupo']
        resultado_df = resultado_df[colunas]
        
        logging.info(f"Análise concluída: {len(resultado_df)} procedimentos em {resultado_df['Grupo'].nunique()} grupos")
        return resultado_df
    else:
        logging.warning("Nenhum procedimento encontrado com ocorrências")
        return pd.DataFrame()

def formatar_planilha(workbook, sheet_name, cnes, nome_hospital, periodo_inicio, periodo_fim, uf="SP"):
    """
    Formata a planilha Excel com o layout desejado
    
    Args:
        workbook: Workbook do openpyxl
        sheet_name: Nome da aba no Excel
        cnes: Código CNES do hospital
        nome_hospital: Nome do hospital
        periodo_inicio: Início do período (MM/YYYY)
        periodo_fim: Fim do período (MM/YYYY)
        uf: UF do hospital
    """
    worksheet = workbook[sheet_name]
    
    # Adicionar cabeçalho
    worksheet.insert_rows(1, 4)  # Insere 4 linhas no topo
    
    # Linha 1: Título principal
    worksheet.cell(row=1, column=1, value=f"Dados detalhados das AIH - {uf}")
    # Linha 2: Subtítulo
    worksheet.cell(row=2, column=1, value="Quantidade aprovada por Procedimento e Ano/mês processamento")
    # Linha 3: Estabelecimento
    worksheet.cell(row=3, column=1, value=f"Estabelecimento: '{cnes} {nome_hospital}'")
    # Linha 4: Período
    worksheet.cell(row=4, column=1, value=f"Periodo: {periodo_inicio}-{periodo_fim}")
    
    # Aplicar formatação ao cabeçalho
    header_font = Font(bold=True, size=12)
    for row in range(1, 5):
        cell = worksheet.cell(row=row, column=1)
        cell.font = header_font
        # Mesclar células nas linhas de cabeçalho
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    
    # Definir nomes das colunas no cabeçalho (linha 5)
    colunas = ['Procedimento']
    for mes in range(1, 13):
        colunas.append(f'Mês {mes}')
    colunas.extend(['Total', 'Grupo'])
    
    # Aplicar nomes das colunas
    for col_idx, nome_coluna in enumerate(colunas, 1):
        worksheet.cell(row=5, column=col_idx, value=nome_coluna)
    
    # Aplicar formatação ao cabeçalho de colunas (linha 5)
    for cell in worksheet[5]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Formatar linhas de dados com cores alternadas por grupo
    grupo_atual = None
    cor_idx = 0
    cores = [
        PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    ]
    
    # Começar na linha 6 (após o cabeçalho)
    for row_idx in range(6, worksheet.max_row + 1):
        # Obter o grupo da última coluna
        grupo = worksheet.cell(row=row_idx, column=worksheet.max_column).value
        
        if grupo != grupo_atual:
            grupo_atual = grupo
            cor_idx = (cor_idx + 1) % 2
        
        # Aplicar cor a todas as células exceto a última (Grupo)
        for col_idx in range(1, worksheet.max_column):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = cores[cor_idx]
            
            # Centralizar colunas numéricas (meses e total)
            if col_idx > 1:  # Todas exceto "Procedimento"
                cell.alignment = Alignment(horizontal='center')
    
    # Ajustar largura das colunas
    worksheet.column_dimensions['A'].width = 60  # Coluna de procedimento
    for col_idx in range(2, worksheet.max_column):  # Meses e Total
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 10
    
    # Ocultar a coluna de Grupo (última coluna)
    last_col_letter = get_column_letter(worksheet.max_column)
    worksheet.column_dimensions[last_col_letter].hidden = True
    
    logging.info(f"Planilha {sheet_name} formatada com sucesso")

def gerar_relatorio_procedimentos(base: str, grupo: str, cnes_list: List[str],
                                 competencia_inicio: str, competencia_fim: str,
                                 anos: List[int], output_dir: str = "."):
    """
    Gera relatórios de procedimentos por hospital e por ano
    
    Args:
        base: SIH ou SIA
        grupo: Código do grupo (SP, RD, etc)
        cnes_list: Lista de CNES
        competencia_inicio: Mês/Ano inicial (MM/YYYY)
        competencia_fim: Mês/Ano final (MM/YYYY)
        anos: Lista de anos para análise
        output_dir: Diretório para salvar os arquivos de saída
    """
    inicio_total = time.time()
    os.makedirs(output_dir, exist_ok=True)
    
    try:
        # Carregar dados
        df = carregar_dados_parquet(base, grupo, cnes_list, competencia_inicio, competencia_fim)
        
        if df.empty:
            logging.error("Não foi possível carregar dados para análise")
            return
        
        # Identificar coluna CNES com mais robustez
        colunas_cnes = [
            col for col in df.columns 
            if 'CNES' in col.upper() and not 
            any(x in col for x in ['PROC', 'ATO', 'CIDPRI', 'PROC'])
        ]
        
        if not colunas_cnes:
            logging.error("Não foi possível identificar a coluna CNES nos dados")
            logging.info(f"Colunas disponíveis: {', '.join(df.columns)}")
            return
            
        # Usar a primeira coluna CNES válida (que não seja parte de nome de procedimento)
        col_cnes = colunas_cnes[0]
        logging.info(f"Usando a coluna '{col_cnes}' como identificadora de estabelecimentos de saúde")
        
        # Padronizar os códigos CNES na coluna
        df[col_cnes] = df[col_cnes].astype(str).apply(lambda x: x.zfill(7) if len(x) <= 7 else x[-7:])
        
        # Valores únicos de CNES para criar um relatório para cada estabelecimento
        cnes_unicos = df[col_cnes].unique()
        logging.info(f"Encontrados {len(cnes_unicos)} estabelecimentos de saúde únicos nos dados")
        
        # Para cada ano solicitado
        for ano in anos:
            inicio_ano = time.time()
            logging.info(f"Processando relatório para o ano {ano}")
            
            nome_arquivo = os.path.join(output_dir, f"Procedimentos Aprovados {ano}.xlsx")
            
            # Criar workbook
            wb = Workbook()
            # Remover a planilha padrão
            ws_default = wb.active
            wb.remove(ws_default)
            
            # Processar cada CNES
            hospitais_processados = 0
            
            # Verificar se os dados têm registros para o ano específico
            # e filtrar por ano antes de processar por CNES
            col_ano = None
            if 'SP_AA' in df.columns:
                col_ano = 'SP_AA'
            elif 'ANO_CMPT' in df.columns:
                col_ano = 'ANO_CMPT'
            else:
                # Procurar qualquer coluna que tenha 'ANO' ou 'AA' no nome
                for col in df.columns:
                    if 'ANO' in col.upper() or 'AA' in col.upper():
                        col_ano = col
                        break
            
            if col_ano:
                try:
                    # Converter para inteiro se necessário
                    if df[col_ano].dtype != 'int64':
                        df[col_ano] = df[col_ano].astype(int)
                        
                    # Filtrar por ano
                    df_ano = df[df[col_ano] == ano]
                    if len(df_ano) == 0:
                        logging.warning(f"Nenhum dado encontrado para o ano {ano}")
                        continue
                    
                    logging.info(f"Encontrados {len(df_ano)} registros para o ano {ano}")
                except Exception as e:
                    logging.error(f"Erro ao filtrar por ano: {e}")
                    df_ano = df  # Usar todos os dados se ocorrer erro
            else:
                df_ano = df  # Usar todos os dados se não encontrar coluna de ano
            
            # Processar cada CNES na lista
            for cnes in cnes_list:
                try:
                    inicio_hospital = time.time()
                    
                    # Normalizar CNES para comparação
                    cnes_limpo = str(cnes).zfill(7) if len(str(cnes)) <= 7 else str(cnes)[-7:]
                    
                    # Filtrar dados para o CNES atual - permitindo correspondência exata
                    df_hospital = df_ano[df_ano[col_cnes] == cnes_limpo].copy()
                    
                    if len(df_hospital) == 0:
                        logging.warning(f"Nenhum dado encontrado para o CNES {cnes} no ano {ano}")
                        continue
                    
                    logging.info(f"Encontrados {len(df_hospital)} registros para o CNES {cnes} no ano {ano}")
                    
                    # Identificar o nome do hospital
                    nome_hospital = obter_nome_hospital(cnes)
                    
                    # Obter a análise de procedimentos para este hospital e ano
                    analise_df = analisar_procedimentos_hospital(df_hospital, ano)
                    
                    if analise_df.empty:
                        logging.warning(f"Nenhum procedimento encontrado para o CNES {cnes} no ano {ano}")
                        continue
                    
                    # Limitar nome da aba a 31 caracteres (limite do Excel)
                    sheet_name = nome_hospital[:31]
                    # Garantir unicidade do nome da aba
                    suffix = 1
                    base_name = sheet_name
                    while sheet_name in wb.sheetnames:
                        sheet_name = f"{base_name[:28]}_{suffix}"
                        suffix += 1
                    
                    # Criar nova aba para este hospital
                    ws = wb.create_sheet(sheet_name)
                    
                    # Adicionar dados à planilha
                    for r_idx, row in enumerate(analise_df.itertuples(index=False), 1):
                        for c_idx, value in enumerate(row, 1):
                            ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Formatar a planilha
                    formatar_planilha(wb, sheet_name, cnes, nome_hospital, 
                                    competencia_inicio, competencia_fim)
                    
                    hospitais_processados += 1
                    logging.info(f"Hospital {nome_hospital} processado em {time.time() - inicio_hospital:.2f} segundos")
                except Exception as e:
                    logging.error(f"Erro ao processar hospital com CNES {cnes}: {str(e)}")
                    import traceback
                    logging.error(traceback.format_exc())
                    continue
            
            # Salvar workbook se tiver planilhas
            if len(wb.sheetnames) > 0:
                wb.save(nome_arquivo)
                logging.info(f"Arquivo {nome_arquivo} salvo com sucesso com {hospitais_processados} hospitais")
            else:
                logging.warning(f"Nenhum dado encontrado para o ano {ano}, arquivo não foi gerado")
            
            logging.info(f"Processamento do ano {ano} concluído em {time.time() - inicio_ano:.2f} segundos")
        
        logging.info(f"Processamento total concluído em {time.time() - inicio_total:.2f} segundos")
        
    except Exception as e:
        logging.error(f"Erro durante a geração de relatórios: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())

def main():
    """
    Função principal - exemplo de uso
    """
    # Configurações
    base = "SIH"  # SIH ou SIA
    grupo = "SP"  # SP, RD, etc.
    cnes_list = [
    "0011800",
    "0012017",
    "2445956",
    "2446030",
    "2447894",
    "2465752",
    "2486199",
    "2546957",
    "2547317",
    "2547783",
    "2550687",
    "2675714",
    "2678179",
    "2709023",
    "6559131",
    "7257406",
    "7530706",
    "7621442",
    "0011738",
    "0011746",
    "0011991",
    "2705591"]  # Lista de CNES ou ["*"] para todos
    competencia_inicio = "01/2024"
    competencia_fim = "12/2024"
    anos = [2024]  # Lista de anos para análise
    output_dir = "relatorios"
    
    # Gerar relatório
    gerar_relatorio_procedimentos(
        base=base,
        grupo=grupo,
        cnes_list=cnes_list,
        competencia_inicio=competencia_inicio,
        competencia_fim=competencia_fim,
        anos=anos,
        output_dir=output_dir
    )

if __name__ == "__main__":
    main() 